from openpyxl import Workbook, load_workbook
from pathlib import Path

def parse_sheet(file_path: Path, sheet_name: str):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                category = cell.split("-")[0]
                data.append((category, cell))
    return data

def get_sheet_data(file_path: Path, sheet_name: str):
    return parse_sheet(file_path, sheet_name)